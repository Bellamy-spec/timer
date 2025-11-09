"""
    作者：郭雨健
    日期：2022.6.12
    功能：图形化界面计时器程序
"""
import tkinter
import time
from tkinter.ttk import Combobox
import winsound
from tkinter import filedialog
import openpyxl
import random
import os
import webbrowser
from xpinyin import Pinyin
import requests
from bs4 import BeautifulSoup
from requests.exceptions import HTTPError, ConnectionError
from urllib.request import urlopen
from PIL import Image
import io
from datetime import datetime
from zhdate import ZhDate

# 浏览器头部信息
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 '
                  '(KHTML, like Gecko) Chrome/102.0.0.0 Safari/537.36'
}

# 星期的对应关系
WEEK_DICT = {0: '星期一',
             1: '星期二',
             2: '星期三',
             3: '星期四',
             4: '星期五',
             5: '星期六',
             6: '星期日'}


def open_img(src):
    """打开图片链接"""
    img_url = urlopen(src, timeout=1).read()
    img_file = io.BytesIO(img_url)
    return Image.open(img_file)


def url_exist(url):
    """判断url是否存在"""
    res = requests.get(url, headers=HEADERS)
    try:
        res.raise_for_status()
    except HTTPError:
        return False
    soup = BeautifulSoup(res.text, 'lxml')
    soups = soup.find_all('div', {'class': 'day7 hide'})
    if len(soups) == 2:
        return True
    else:
        return False


def write8(st, row, start_col, message_list):
    """一次写入8个数据"""
    for i in range(8):
        st.cell(row=row, column=start_col + i).value = message_list[i]


def write40(soup, st, start_col):
    """处理8小时的天气数据"""
    ele_list = soup.find_all('li')
    text_list = []
    for ele in ele_list:
        text_list.append(ele.text)
    text_list = text_list[8:]

    # 写入时间，天气，温度，风向，风力
    for i in range(5):
        row = i + 1
        start = 8 * (i - 1)
        write8(st, row, start_col, text_list[start:])


def hours_weather(url, st):
    """处理24小时天气信息"""
    # 读取网页内容并解析
    res = requests.get(url, headers=HEADERS)
    res.raise_for_status()
    soup = BeautifulSoup(res.text, 'lxml')

    # 读取并处理24小时天气数据
    soups = [soup.find('div', {'class': 'day7'})]
    soups += soup.find_all('div', {'class': 'day7 hide'})
    for i in range(3):
        write40(soups[i], st, 8 * i + 1)


def month_weather(url, st):
    """处理30天天气信息"""
    # 写入表头
    st['A1'].value = '日期'
    st['B1'].value = '星期'
    st['C1'].value = '天气'
    st['D1'].value = '温度'
    st['E1'].value = '空气质量'
    st['F1'].value = '风向风力'

    # 读取网页内容并解析
    res = requests.get(url, headers=HEADERS)
    res.raise_for_status()

    # 去注释
    tt = res.text.replace('<!-- ', '')
    tt = tt.replace(' -->', '')
    soup = BeautifulSoup(tt, 'lxml')

    # 找到每天天气数据
    soup = soup.find('ul', {'class': 'weaul'})
    soups = soup.find_all('li')

    # 起始行
    row = 2

    # 循环遍历每个元素写入每一天
    for sp in soups:
        # 日期
        date = sp.find('span', {'class': 'fl'}).text
        st.cell(row=row, column=1).value = date

        # 星期
        week = sp.find('span', {'class': 'fr'}).text
        st.cell(row=row, column=2).value = week

        # 天气
        wea = sp.find('div', {'class': 'weaul_z'}).text
        st.cell(row=row, column=3).value = wea

        # 温度
        tem = sp.find_all('div', {'class': 'weaul_z'})[1].text
        st.cell(row=row, column=4).value = tem

        # 空气质量
        air = sp.find('div', {'class': 'weaul_w'}).text
        st.cell(row=row, column=5).value = air

        # 风向风力
        wind = sp.find('div', {'class': 'weaul_s'}).text
        st.cell(row=row, column=6).value = wind

        row += 1


def city_to_url(city):
    """对给定城市生成天气网url"""
    # 取得城市的拼音
    p = Pinyin()
    r = p.get_pinyin(city)
    city_py = ''.join(r.split('-'))

    # 取得城市天气网页URL
    weather_url = 'https://www.tianqi.com/' + city_py + '/'

    return weather_url


def new_wind(msg, rt):
    """弹出一个提示信息的新窗体"""
    nw = tkinter.Toplevel(rt)
    nw.geometry('500x200')
    nw.title('提示信息')
    lb = tkinter.Label(nw, text=msg + '！', font=('黑体', 14))
    lb.place(relx=0.1, rely=0.2,
             relwidth=0.8, relheight=0.5)

    # 退出按钮
    exit_button = tkinter.Button(nw, text='确定', font=('黑体', 14),
                                 command=nw.destroy)
    exit_button.place(relx=0.45, relwidth=0.1,
                      rely=0.75, relheight=0.15)


def two_digitals(s):
    """数字字符串作两位化处理"""
    if len(s) == 1:
        # 只有一位时，前面自动补0
        s = '0' + s
    elif len(s) > 2:
        # 大于两位时，只取后两位
        s = s[-2:]
    return s


def open_baidu():
    """直接打开百度"""
    webbrowser.open('https://www.baidu.com')


class AppTk(tkinter.Tk):
    """包含应用的窗体类"""

    def __init__(self):
        super().__init__()

        # 窗体上的应用
        self.app = AddTime(self)

        # 启动秒表程序的按钮
        self.add_button = tkinter.Button(self, text='秒表',
                                         fg='black',
                                         font=('黑体', 12),
                                         command=self.create_app1)
        self.add_button.place(relx=0, rely=0,
                              relwidth=0.125, relheight=0.0625)

        # 启动倒计时程序的按钮
        self.minus_button = tkinter.Button(self, text='倒计时',
                                           fg='black',
                                           font=('黑体', 12),
                                           command=self.create_app2)
        self.minus_button.place(relx=0.125, rely=0,
                                relwidth=0.125, relheight=0.0625)

        # 启动点名程序的按钮
        self.roll_button = tkinter.Button(self, text='点名',
                                          fg='black',
                                          font=('黑体', 12),
                                          command=self.create_app3)
        self.roll_button.place(relx=0.25, rely=0,
                               relwidth=0.125, relheight=0.0625)

        # 启动打开文件程序的按钮
        self.open_button = tkinter.Button(self, text='打开',
                                          fg='black',
                                          font=('黑体', 12),
                                          command=self.create_app4)
        self.open_button.place(relx=0.375, rely=0,
                               relwidth=0.125, relheight=0.0625)

        # 启动天气程序的按钮
        self.weather_button = tkinter.Button(self, text='天气',
                                             fg='black',
                                             font=('黑体', 12),
                                             command=self.create_app5)
        self.weather_button.place(relx=0.5, rely=0,
                                  relwidth=0.125, relheight=0.0625)

        # 启动时钟程序的按钮
        self.clock_button = tkinter.Button(self, text='时钟',
                                           fg='black',
                                           font=('黑体', 12),
                                           command=self.create_app6)
        self.clock_button.place(relx=0.625, rely=0,
                                relwidth=0.125, relheight=0.0625)

        # 已打开的应用
        self.opened_apps = {'add': self.app}

    def create_app1(self):
        """创建秒表应用程序"""
        # 隐藏原有程序的所有控件
        self.app.hide()

        if 'add' in self.opened_apps.keys():
            # 创建过的应用程序重新显示出来即可
            self.app = self.opened_apps['add']
            self.app.show()
        else:
            # 创建新的应用程序
            self.app = AddTime(self)
            self.opened_apps['add'] = self.app

    def create_app2(self):
        """创建倒计时应用程序"""
        # 隐藏原有程序的所有控件
        self.app.hide()

        if 'minus' in self.opened_apps.keys():
            # 创建过的应用程序重新显示出来即可
            self.app = self.opened_apps['minus']
            self.app.show()
        else:
            # 创建新的应用程序
            self.app = MinusTime(self)
            self.opened_apps['minus'] = self.app

    def create_app3(self):
        """创建点名应用程序"""
        # 隐藏原有程序的所有控件
        self.app.hide()

        if 'name' in self.opened_apps.keys():
            # 创建过的应用程序重新显示出来即可
            self.app = self.opened_apps['name']
            self.app.show()
        else:
            # 创建新的应用程序
            self.app = RollName(self)
            self.opened_apps['name'] = self.app

    def create_app4(self):
        """创建打开文件的应用程序"""
        # 隐藏原有程序的所有控件
        self.app.hide()

        if 'open' in self.opened_apps.keys():
            # 创建过的应用程序重新显示出来即可
            self.app = self.opened_apps['open']
            self.app.show()
        else:
            # 创建新的应用程序
            self.app = SimpleOpen(self)
            self.opened_apps['open'] = self.app

    def create_app5(self):
        """创建查询天气的app"""
        # 隐藏原有程序的所有控件
        self.app.hide()

        if 'weather' in self.opened_apps.keys():
            # 创建过的应用程序重新显示出来即可
            self.app = self.opened_apps['weather']
            self.app.show()
        else:
            # 创建新的应用程序
            self.app = WeatherConsult(self)
            self.opened_apps['weather'] = self.app

    def create_app6(self):
        """创建时钟App"""
        # 隐藏原有程序的控件
        self.app.hide()

        if 'clock' in self.opened_apps.keys():
            # 创建过的应用程序显示出来即可
            self.app = self.opened_apps['clock']
            self.app.show()
        else:
            # 创建新的应用程序
            self.app = Clock(self)
            self.opened_apps['clock'] = self.app


class Clock:
    """时钟程序"""

    def __init__(self, rt):
        """初始化程序部件"""
        # 显示时间的标签
        self.time_label = tkinter.Label(rt, text='00:00:00',
                                        fg='black',
                                        font=('Arial', 50))

        # 显示日期的标签
        self.date_label = tkinter.Label(rt, text='20XX年XX月XX日，星期X',
                                        fg='#556677',
                                        font=('黑体', 20))

        # 显示农历日期的标签
        self.lunar_label = tkinter.Label(rt, text='农历XXXX年XX月XX日',
                                         fg='#776655',
                                         font=('黑体', 16))

        # 部件组
        self.units = [self.time_label, self.date_label, self.lunar_label]

        self.show()
        self.update_time()

    def update_time(self):
        """更新时间"""
        # 获取当前时间
        dt = datetime.now()

        # 更新时间显示
        h = two_digitals(str(dt.hour))
        m = two_digitals(str(dt.minute))
        s = two_digitals(str(dt.second))
        time_str = '{}:{}:{}'.format(h, m, s)
        self.time_label.config(text=time_str)

        # 更新日期显示
        wk = WEEK_DICT[dt.weekday()]
        date_str = '{}年{}月{}日，{}'.format(dt.year, dt.month, dt.day, wk)
        self.date_label.config(text=date_str)

        # 更新农历日期显示
        lunar_date_str = ZhDate.from_datetime(dt).__str__()
        self.lunar_label.config(text=lunar_date_str)

        # 持续更新
        self.date_label.after(100, self.update_time)

    def show(self):
        """显示部件"""
        self.time_label.place(relx=0.2, rely=0.3,
                              relwidth=0.6, relheight=0.2)
        self.date_label.place(relx=0.2, rely=0.5,
                              relwidth=0.6, relheight=0.2)
        self.lunar_label.place(relx=0.2, relwidth=0.6,
                               rely=0.7, relheight=0.15)

    def hide(self):
        """隐藏部件"""
        for unit in self.units:
            unit.place_forget()


class AddTime:
    """正向计时程序"""

    def __init__(self, rt):
        """初始化程序部件"""
        # 控制按钮
        self.command_button = tkinter.Button(rt, text='开始',
                                             bg='white',
                                             fg='black',
                                             font=('黑体', 16),
                                             command=self.start_or_pause)

        # 重置按钮
        self.reset_button = tkinter.Button(rt, text='重置',
                                           bg='white',
                                           fg='black',
                                           font=('黑体', 16),
                                           command=self.reset)

        # 显示时间的标签
        self.time_label = tkinter.Label(rt, text='00:00:00',
                                        fg='black',
                                        font=('Arial', 50))

        # 部件组
        self.units = [self.command_button,
                      self.reset_button,
                      self.time_label]
        self.show()

        # 活跃状态
        self.active = False
        self.counting = False

        # 初始化开始时间
        self.start_time = time.time()
        self.pause_start = None
        self.paused = 0

    def update_time(self):
        """实时更新时间"""
        ns = round(time.time() - self.start_time - self.paused, 2)

        # 后两位，中间两位，前两位
        if ns < 3600:
            p = str(int(ns * 100 % 100))
            s = str(int(ns) % 60)
            m = str(int(ns) // 60)
        else:
            p = str(int(ns) % 60)
            s = str(int(ns) // 60 % 60)
            m = str(int(ns) // 3600)

        # 两位化处理
        msgs = []
        for i in [m, s, p]:
            msgs.append(two_digitals(i))

        # 生成显示信息
        if ns < 3600:
            msg = '{}:{}.{}'.format(*msgs)
        else:
            msg = '{}:{}:{}'.format(*msgs)

        # 重新显示
        self.time_label.config(text=msg)

        # 活跃状态下循环
        if self.active:
            self.time_label.after(10, self.update_time)
        else:
            if not self.counting:
                self.time_label.config(text='00:00:00')

    def start_or_pause(self):
        """开始计时或暂停计时"""
        if not self.counting:
            # 没有在计时，则重置开始时间
            self.start_time = time.time()
            self.counting = True
        else:
            if not self.active:
                self.paused += time.time() - self.pause_start

        if self.active:
            # 活跃状态下暂停计时
            self.active = False
            self.pause_start = time.time()
            self.command_button.config(text='开始')
        else:
            # 非活跃状态下开始计时
            self.active = True
            self.command_button.config(text='暂停')
            self.update_time()

    def reset(self):
        """重置"""
        self.active = False
        self.counting = False
        self.time_label.config(text='00:00:00')
        self.paused = 0
        self.command_button.config(text='开始')

    def hide(self):
        """隐藏部件"""
        for unit in self.units:
            unit.place_forget()

    def show(self):
        """显示部件"""
        self.command_button.place(relx=0.28, rely=0.7,
                                  relwidth=0.2, relheight=0.12)
        self.reset_button.place(relx=0.52, rely=0.7,
                                relwidth=0.2, relheight=0.12)
        self.time_label.place(relx=0.2, rely=0.3,
                              relwidth=0.6, relheight=0.2)


class MinusTime:
    """倒计时程序"""

    def __init__(self, rt):
        """初始化程序部件"""
        # 根窗体
        self.rt = rt

        # 控制按钮
        self.command_button = tkinter.Button(rt, text='开始',
                                             bg='white',
                                             fg='black',
                                             font=('黑体', 16),
                                             command=self.start_or_pause)

        # 重置按钮
        self.reset_button = tkinter.Button(rt, text='重置',
                                           bg='white',
                                           fg='black',
                                           font=('黑体', 16),
                                           command=self.reset)

        # 显示时间的标签
        self.time_label = tkinter.Label(rt, text='00:00:00',
                                        fg='black',
                                        font=('Arial', 50))

        # 用于输入时间的部件
        self.msg_label0 = tkinter.Label(rt, text='请设置倒计时时间：',
                                        fg='black',
                                        font=('黑体', 12))
        self.comma1 = tkinter.Label(rt, text=':',
                                    fg='black',
                                    font=('Arial', 50))
        self.comma2 = tkinter.Label(rt, text=':',
                                    fg='black',
                                    font=('Arial', 50))
        self.hour_input = Combobox(rt, font=('Arial', 48),
                                   values=list(range(100)))
        self.minute_input = Combobox(rt, font=('Arial', 48),
                                     values=list(range(60)))
        self.second_input = Combobox(rt, font=('Arial', 48),
                                     values=list(range(60)))

        # 输入部件组
        self.input_units = [self.msg_label0,
                            self.comma1,
                            self.comma2,
                            self.hour_input,
                            self.minute_input,
                            self.second_input]

        # 设置铃声的按钮
        self.set_ring = tkinter.Button(rt,
                                       text='点此设置铃声\n(仅支持wav格式)',
                                       fg='black',
                                       font=('黑体', 10),
                                       command=self.set_music)

        # 部件组
        self.units = [self.command_button,
                      self.reset_button,
                      self.time_label,
                      self.set_ring] + self.input_units

        # 活跃状态
        self.active = False
        self.counting = False
        self.show()

        # 初始化开始时间
        self.start_time = time.time()
        self.pause_start = None
        self.paused = 0
        self.total = 0
        self.time_up = False

        # 默认铃声
        self.music = '14060.wav'

    def set_music(self):
        """弹出文件对话框设置铃声"""
        filename = filedialog.askopenfilename()
        if filename:
            if filename.split('.')[-1] == 'wav':
                self.music = filename
                new_wind('设置成功', self.rt)
            else:
                new_wind('设置失败', self.rt)
        else:
            new_wind('设置失败', self.rt)

    def show_input(self):
        """显示输入控件"""
        self.hour_input.place(relx=0.2, rely=0.3,
                              relwidth=0.16, relheight=0.2)
        self.minute_input.place(relx=0.42, rely=0.3,
                                relwidth=0.16, relheight=0.2)
        self.second_input.place(relx=0.64, rely=0.3,
                                relwidth=0.16, relheight=0.2)
        self.comma1.place(relx=0.36, rely=0.3,
                          relwidth=0.06, relheight=0.2)
        self.comma2.place(relx=0.58, rely=0.3,
                          relwidth=0.06, relheight=0.2)
        self.msg_label0.place(relx=0.35, rely=0.15,
                              relwidth=0.3, relheight=0.0625)

    def update(self):
        """实时更新显示"""
        # 已过去的时间
        ns = round(time.time() - self.start_time - self.paused, 2)

        # 剩余时间
        ns = self.total - ns

        if ns > 0:
            # 后两位，中间两位，前两位
            if ns < 3600:
                p = str(int(ns * 100 % 100))
                s = str(int(ns) % 60)
                m = str(int(ns) // 60)
            else:
                p = str(int(ns) % 60)
                s = str(int(ns) // 60 % 60)
                m = str(int(ns) // 3600)

            # 两位化处理
            msgs = []
            for i in [m, s, p]:
                msgs.append(two_digitals(i))

            # 生成显示信息
            if ns < 3600:
                msg = '{}:{}.{}'.format(*msgs)
            else:
                msg = '{}:{}:{}'.format(*msgs)

            # 重新显示
            self.time_label.config(text=msg)

            # 活跃状态下循环
            if self.active:
                self.time_label.after(10, self.update)
            else:
                if not self.counting:
                    self.time_label.config(text='00:00:00')
        else:
            # 时间到
            self.ring()

    def start_or_pause(self):
        """开始或暂停"""
        if self.time_up:
            # 已经时间到，不执行任何操作
            return

        if not self.counting:
            # 读取输入组件内容并计算总秒数
            try:
                if self.hour_input.get():
                    h = int(self.hour_input.get())
                else:
                    h = 0
                if self.minute_input.get():
                    m = int(self.minute_input.get())
                else:
                    m = 0
                if self.second_input.get():
                    s = int(self.second_input.get())
                else:
                    s = 0
            except ValueError:
                # 处理输入格式错误
                new_wind('请输入正确的时间格式', self.rt)
                return
            else:
                self.total = h * 3600 + m * 60 + s
                if self.total >= 360000 or self.total < 0:
                    # 处理出界
                    new_wind('超出程序能够处理的范围', self.rt)
                    return

            # 隐藏输入控件
            for unit in self.input_units:
                unit.place_forget()

            # 显示倒计时标签
            self.time_label.place(relx=0.2, rely=0.3,
                                  relwidth=0.6, relheight=0.2)

            # 记录开始时间
            self.start_time = time.time()

            # 开始
            self.counting = True
        else:
            if not self.active:
                self.paused += time.time() - self.pause_start

        if self.active:
            # 活跃状态下暂停计时
            self.active = False
            self.pause_start = time.time()
            self.command_button.config(text='继续')
        else:
            # 非活跃状态下开始计时
            self.active = True
            self.command_button.config(text='暂停')
            self.update()

    def reset(self):
        """重置"""
        # 关闭音乐
        winsound.PlaySound(None, winsound.SND_PURGE)

        # 关闭三开关
        self.counting = False
        self.active = False
        self.time_up = False

        # 隐藏时间面板
        self.time_label.place_forget()

        # 显示输入面板
        self.show_input()

        # 重置按钮信息
        self.command_button.config(text='开始')

        # 重置暂停时间
        self.paused = 0

    def ring(self):
        """提醒时间到"""
        # 开关调整
        self.active = False
        self.time_up = True

        self.time_label['text'] = '时间到！'

        # 响铃
        winsound.PlaySound(self.music, winsound.SND_ASYNC)

    def hide(self):
        """隐藏部件"""
        for unit in self.units:
            unit.place_forget()

    def show(self):
        """显示部件"""
        self.command_button.place(relx=0.28, rely=0.7,
                                  relwidth=0.2, relheight=0.12)
        self.reset_button.place(relx=0.52, rely=0.7,
                                relwidth=0.2, relheight=0.12)
        self.set_ring.place(relx=0.8, rely=0.88,
                            relwidth=0.2, relheight=0.12)
        if self.counting:
            # 显示倒计时标签
            self.time_label.place(relx=0.2, rely=0.3,
                                  relwidth=0.6, relheight=0.2)
        else:
            self.show_input()


class RollName:
    """点名应用程序"""

    def __init__(self, rt):
        """初始化程序部件"""
        # 根窗体
        self.rt = rt

        # 录入姓名的按钮
        self.add_button = tkinter.Button(rt, text='录入',
                                         bg='white',
                                         fg='black',
                                         font=('黑体', 16),
                                         command=self.tips)

        # 控制开始与停止的按钮
        self.command_button = tkinter.Button(rt, text='开始',
                                             bg='white',
                                             fg='black',
                                             font=('黑体', 16),
                                             command=self.start_or_stop)

        # 显示信息的标签
        self.msg_label = tkinter.Label(rt, text='这里将随机生成滚动信息',
                                       fg='black',
                                       font=('黑体', 32))

        # 控件组
        self.units = [self.add_button,
                      self.command_button,
                      self.msg_label]
        self.show()

        # 存储姓名信息的列表
        self.name_list = []

        # 活跃状态
        self.active = False

    def tips(self):
        """录入前提示操作"""
        message = '请先把所有要录入的信息放入同一个Excel表格的A列，' \
                  '注意该Excel表格文件名后缀必须为.xlsx，点击“下一' \
                  '步”以选择该文件'
        nw = tkinter.Toplevel(self.rt)
        nw.geometry('500x200')
        nw.title('操作提示')
        lb = tkinter.Label(nw, text=message, font=('黑体', 12),
                           wraplength=400, justify='left')
        lb.place(relx=0.1, rely=0.2,
                 relwidth=0.8, relheight=0.5)

        # 继续操作的按钮
        next_button = tkinter.Button(nw, text='下一步', font=('黑体', 12),
                                     command=lambda: self.add_name(nw))
        next_button.place(relx=0.42, rely=0.75,
                          relwidth=0.16, relheight=0.15)

    def add_name(self, nw):
        """录入姓名"""
        # 先销毁上一步提示操作的窗口
        nw.destroy()

        # 弹出文件对话框选择文件
        filename = filedialog.askopenfilename()
        if filename:
            if filename.split('.')[-1] == 'xlsx':
                # 读取名单并录入
                wb = openpyxl.load_workbook(filename)
                st = wb.active
                for row in range(1, st.max_row + 1):
                    cell = 'A' + str(row)
                    name = st[cell].value
                    self.name_list.append(name)

                # 随手关闭文件是个好习惯
                wb.close()

                # 提示成功
                new_wind('录入成功', self.rt)
            else:
                new_wind('请选择xlsx格式的文件', self.rt)
        else:
            new_wind('你没有选择任何文件', self.rt)

    def start_or_stop(self):
        """开始或停止"""
        if self.active:
            # 活跃状态下停止
            self.active = False
            self.command_button.config(text='开始')
        else:
            # 非活跃状态下开始
            if self.name_list:
                self.active = True
                self.command_button.config(text='停止')
                self.roll()
            else:
                new_wind('你还没有录入信息', self.rt)

    def roll(self):
        """滚动起来！"""
        name = random.choice(self.name_list)
        self.msg_label.config(text=name)
        if self.active:
            self.msg_label.after(50, self.roll)

    def hide(self):
        """隐藏部件"""
        for unit in self.units:
            unit.place_forget()

    def show(self):
        """显示部件"""
        self.add_button.place(relx=0.28, rely=0.7,
                              relwidth=0.2, relheight=0.12)
        self.command_button.place(relx=0.52, rely=0.7,
                                  relwidth=0.2, relheight=0.12)
        self.msg_label.place(relx=0.1, rely=0.3,
                             relwidth=0.8, relheight=0.2)


class WeatherConsult:
    """查询天气的程序"""

    def __init__(self, rt):
        """初始化程序部件"""
        # 根窗体
        self.rt = rt

        # 用于输入城市的文本框
        self.input_entry = tkinter.Entry(rt, font=('黑体', 16))

        # 提示信息的标签
        self.msg_label = tkinter.Label(rt, text='请输入你要查询的城市：',
                                       font=('黑体', 16))

        # 1.0版本直接打开天气网站的按钮
        self.button1 = tkinter.Button(rt, text='打开天气网页',
                                      font=('宋体', 14),
                                      bg='#EEFFFF',
                                      command=self.open_url)

        # 2.0版本显示天气信息的按钮
        self.button2 = tkinter.Button(rt, text='显示天气信息',
                                      font=('宋体', 14),
                                      bg='#FFEEFF',
                                      command=self.output_weather)

        # 4.0版本导出天气数据
        self.button4 = tkinter.Button(rt, text='导出天气数据',
                                      font=('宋体', 14),
                                      bg='#FFFFEE',
                                      command=self.save_weather)

        # 显示输出信息的标签
        self.output_label = tkinter.Label(rt, text='这里将显示天气信息',
                                          font=('黑体', 28))

        # 部件组
        self.units = [self.input_entry,
                      self.msg_label,
                      self.button1,
                      self.button2,
                      self.button4,
                      self.output_label]
        self.show()

    def hide(self):
        """隐藏部件"""
        for unit in self.units:
            unit.place_forget()

    def show(self):
        """显示部件"""
        self.input_entry.place(relx=0.6, relwidth=0.2,
                               rely=0.125, relheight=0.075)
        self.msg_label.place(relx=0.2, relwidth=0.4,
                             rely=0.125, relheight=0.075)
        self.button1.place(relx=0.125, relwidth=0.25,
                           rely=0.25, relheight=0.1)
        self.button2.place(relx=0.375, relwidth=0.25,
                           rely=0.25, relheight=0.1)
        self.button4.place(relx=0.625, relwidth=0.25,
                           rely=0.25, relheight=0.1)
        self.output_label.place(relx=0.1, relwidth=0.8,
                                rely=0.4, relheight=0.5)

    def output_weather(self):
        """输出天气信息"""
        city = self.input_entry.get()
        if city:
            # TODO: 爬虫取得该城市天气信息并输出
            weather_url = city_to_url(city)

            try:
                if url_exist(weather_url + 'today/'):
                    # 读取网页源代码
                    res = requests.get(weather_url, headers=HEADERS)
                    res.raise_for_status()
                    tt = res.text

                    # bs4解析网页
                    soup = BeautifulSoup(tt, 'lxml')

                    # 重新取得准确的城市名称
                    city_soup = soup.find('dd', {'class': 'name'})
                    city = city_soup.text[:-8]

                    # 解析天气数据
                    weather_soup1 = soup.find('dd', {'class': 'weather'})

                    # # 取得天气图像
                    # img_soup = weather_soup1.find('img')
                    # weather_img_url = 'https:' + img_soup.get('src')
                    # requests.get(weather_img_url).raise_for_status()
                    # open_img(weather_img_url)

                    # 取得当前气温
                    tem_now_soup = weather_soup1.find('p', {'class': 'now'})
                    tem_now = tem_now_soup.find('b').text

                    # 取得当天气温
                    tem_soup = weather_soup1.find('span')
                    tem = tem_soup.text

                    # 取得当天天气状况
                    wtr = tem_soup.find('b').text

                    # 取得纯净的气温数据
                    tem = tem.replace(wtr, '')

                    # 生成天气信息字符串
                    weather_message = '{}今天{}，{}，当前温度为{}℃'.format(city, wtr, tem, tem_now)

                    # 解析湿度，风力风向，紫外线数据
                    weather_soup2 = soup.find('dd', {'class': 'shidu'})
                    shidu_list = weather_soup2.find_all('b')
                    bt_list = []
                    for shidu in shidu_list:
                        bt_list.append(shidu.text)
                    shidu_str = ' '.join(bt_list)

                    # 解析空气数据
                    weather_soup3 = soup.find('dd', {'class': 'kongqi'})
                    kongqi_str = weather_soup3.find('h5').text + ' '
                    kongqi_str += weather_soup3.find('h6').text

                    # 输出天气
                    output_message = '\n'.join([weather_message, shidu_str, kongqi_str])
                    self.output_label.config(text=output_message, font=('宋体', 16),
                                             justify='left')
                else:
                    new_wind('您输入的城市或不存在', self.rt)
            except ConnectionError:
                new_wind('请先检查网络连接', self.rt)
        else:
            new_wind('请先输入城市', self.rt)

    def open_url(self):
        """打开天气网页"""
        city_url = city_to_url(self.input_entry.get())
        webbrowser.open(city_url)

    def save_weather(self):
        """天气结果保存为文件"""
        city = self.input_entry.get()
        if city:
            # 取得城市天气url
            weather_url = city_to_url(city)

            # 24小时天气url
            hours_url = weather_url + 'today/'

            # 30天天气url
            month_url = weather_url + '30/'

            # 创建好用来存储天气信息的表格
            wb = openpyxl.Workbook()
            st1 = wb.create_sheet('24小时天气')
            st2 = wb.create_sheet('30天天气')
            del wb['Sheet']

            # 写入天气
            try:
                hours_weather(hours_url, st1)
                month_weather(month_url, st2)
            except IndexError:
                new_wind('您输入的城市或不存在', self.rt)
            except ConnectionError:
                new_wind('请先检查网络连接', self.rt)
            else:
                # 保存文件
                fn = city + '天气.xlsx'
                base_path = filedialog.askdirectory()
                if base_path:
                    fp = os.path.join(base_path, fn)
                    wb.save(fp)
                    new_wind('保存成功', self.rt)
                else:
                    new_wind('您没有选择文件夹', self.rt)

            # 关闭文件
            wb.close()
        else:
            new_wind('请先输入城市', self.rt)


class SimpleOpen:
    """打开任意文件的程序"""

    def __init__(self, rt):
        """初始化程序部件"""
        # 选择并且打开文件的按钮
        self.open_button = tkinter.Button(rt, text='选择一个文件\n并且直接打开',
                                          bg='white',
                                          fg='black',
                                          font=('黑体', 32),
                                          command=self.open_file)

        # 直接打开百度的按钮
        self.baidu_button = tkinter.Button(rt, text='打开百度',
                                           bg='white',
                                           fg='black',
                                           font=('黑体', 16),
                                           command=open_baidu)

        # 部件组
        self.units = [self.open_button, self.baidu_button]
        self.show()

    @staticmethod
    def open_file():
        """选择并打开文件"""
        filename = filedialog.askopenfilename()
        os.system(filename)
        os.system('exit')

    def hide(self):
        """隐藏部件"""
        for unit in self.units:
            unit.place_forget()

    def show(self):
        """显示部件"""
        self.open_button.place(relx=0.15, rely=0.2,
                               relwidth=0.7, relheight=0.4)
        self.baidu_button.place(relx=0.4, rely=0.7,
                                relwidth=0.2, relheight=0.12)


def main():
    """主函数"""
    # 主窗体标题，大小，背景色，主循环
    window = AppTk()
    window.title('计时器')
    window.geometry('600x400')
    window.config(bg='#EEEEEE')

    # 设置窗体尺寸为不可变
    window.minsize(600, 400)
    window.maxsize(600, 400)

    window.mainloop()


if __name__ == '__main__':
    main()
